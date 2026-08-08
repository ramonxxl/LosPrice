"""
LosPrice - Controller de relatorios
====================================

Gera PDF (ReportLab) e Excel (OpenPyXL), alem da engenharia de cardapio.

A ficha tecnica em PDF e o item mais importante daqui: e o que garante
que o custo calculado no sistema seja o custo real da cozinha. Sem
gramatura padronizada na parede, o pastel de hoje nao e o de ontem.
"""

import os
from datetime import datetime

from core.calculo import (
    ABACAXI, ENIGMA, ESTRELA, PUXADOR, ROTULO_BASE, classificar_cardapio,
)
from database.conexao import PASTA_RAIZ, conectar, obter_config
from utils.tema import formatar_moeda, formatar_moeda_precisa, formatar_pct

PASTA_RELATORIOS = os.path.join(PASTA_RAIZ, "relatorios")

# Paleta usada nos PDFs (hex, convertido na hora de desenhar)
LARANJA = "#FF6B00"
VERDE = "#16A34A"
VERMELHO = "#DC2626"
CINZA = "#6B7280"
CINZA_CLARO = "#F0F2F5"
PRETO = "#111827"

CLASSIFICACAO = {
    ESTRELA: ("Vende muito e lucra muito", "Destaque no cardapio e nas fotos.", VERDE),
    PUXADOR: ("Vende muito e lucra pouco", "Suba o preco aos poucos ou barateie a ficha.", "#D97706"),
    ENIGMA: ("Lucra muito e vende pouco", "Divulgue: sugestao do dia, combo, destaque.", "#2563EB"),
    ABACAXI: ("Nao vende e nao lucra", "Candidato a sair do cardapio.", VERMELHO),
}


def garantir_pasta():
    os.makedirs(PASTA_RELATORIOS, exist_ok=True)
    return PASTA_RELATORIOS


def _nome_arquivo(prefixo, extensao):
    carimbo = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return os.path.join(garantir_pasta(), f"{prefixo}_{carimbo}.{extensao}")


def _empresa():
    return obter_config("empresa_nome", "") or "LosPrice"


def abrir_arquivo(caminho):
    """Abre no aplicativo padrao do sistema."""
    try:
        os.startfile(caminho)
        return True
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Engenharia de cardapio
# ---------------------------------------------------------------------------


def engenharia():
    """
    Cruza margem x popularidade e classifica cada produto.
    Precisa do volume mensal informado pelo usuario.
    """
    with conectar() as cur:
        cur.execute(
            """
            SELECT r.id, r.nome, r.categoria, r.custo_unitario, r.vendas_mes,
                   AVG(p.margem_real_pct) AS margem,
                   AVG(p.lucro_liquido)   AS lucro,
                   AVG(p.preco_praticado) AS preco
              FROM receitas r
              LEFT JOIN precificacao p ON p.receita_id = r.id
             WHERE r.ativo = 1
             GROUP BY r.id
             ORDER BY r.nome COLLATE NOCASE
            """
        )
        produtos = [dict(l) for l in cur.fetchall()]

    precificados = [p for p in produtos if p["margem"] is not None]
    if not precificados:
        return {"produtos": produtos, "classificados": [], "pronto": False,
                "margem_media": 0, "vendas_media": 0}

    margem_media = sum(p["margem"] for p in precificados) / len(precificados)
    com_vendas = [p for p in precificados if (p["vendas_mes"] or 0) > 0]

    if not com_vendas:
        return {"produtos": produtos, "classificados": [], "pronto": False,
                "margem_media": margem_media, "vendas_media": 0}

    vendas_media = sum(p["vendas_mes"] for p in com_vendas) / len(com_vendas)

    # Corte de popularidade pelo metodo Kasavana-Smith: 70% da media, nao a
    # media cheia. Sem isso um unico item de giro muito baixo derruba a
    # referencia e produtos saudaveis aparecem como "Abacaxi".
    vendas_corte = vendas_media * 0.70

    for produto in com_vendas:
        produto["classe"] = classificar_cardapio(
            produto["margem"], produto["vendas_mes"], margem_media, vendas_corte)
        produto["lucro_mes"] = (produto["lucro"] or 0) * produto["vendas_mes"]
        texto, acao, cor = CLASSIFICACAO[produto["classe"]]
        produto["classe_texto"] = texto
        produto["classe_acao"] = acao
        produto["classe_cor"] = cor

    com_vendas.sort(key=lambda p: p["lucro_mes"], reverse=True)

    return {
        "produtos": produtos,
        "classificados": com_vendas,
        "pronto": True,
        "margem_media": margem_media,
        "vendas_media": vendas_media,
        "vendas_corte": vendas_corte,
        "lucro_mes_total": sum(p["lucro_mes"] for p in com_vendas),
    }


def salvar_vendas(receita_id, vendas_mes):
    with conectar() as cur:
        cur.execute("UPDATE receitas SET vendas_mes = ? WHERE id = ?",
                    (max(float(vendas_mes or 0), 0), receita_id))


# ---------------------------------------------------------------------------
# PDF - base
# ---------------------------------------------------------------------------


def _estilos():
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.colors import HexColor

    base = getSampleStyleSheet()
    return {
        "titulo": ParagraphStyle("titulo", parent=base["Title"], fontSize=20,
                                 textColor=HexColor(PRETO), spaceAfter=2,
                                 alignment=0),
        "subtitulo": ParagraphStyle("subtitulo", parent=base["Normal"], fontSize=9,
                                    textColor=HexColor(CINZA), spaceAfter=14),
        "secao": ParagraphStyle("secao", parent=base["Heading2"], fontSize=11,
                                textColor=HexColor(LARANJA), spaceBefore=14,
                                spaceAfter=6),
        "corpo": ParagraphStyle("corpo", parent=base["Normal"], fontSize=9,
                                textColor=HexColor(PRETO), leading=13),
        "nota": ParagraphStyle("nota", parent=base["Normal"], fontSize=7.5,
                               textColor=HexColor(CINZA), leading=10),
    }


def _cabecalho(elementos, estilos, titulo, subtitulo):
    from reportlab.platypus import Paragraph

    elementos.append(Paragraph(titulo, estilos["titulo"]))
    elementos.append(Paragraph(subtitulo, estilos["subtitulo"]))


def _rodape(canvas, documento):
    from reportlab.lib.colors import HexColor
    from reportlab.lib.units import mm

    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(HexColor(CINZA))
    canvas.drawString(
        18 * mm, 12 * mm,
        f"{_empresa()}  ·  gerado pelo LosPrice em "
        f"{datetime.now().strftime('%d/%m/%Y as %H:%M')}")
    canvas.drawRightString(canvas._pagesize[0] - 18 * mm, 12 * mm,
                           f"pagina {documento.page}")
    canvas.restoreState()


def _documento(caminho, paisagem=False):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate

    return SimpleDocTemplate(
        caminho,
        pagesize=landscape(A4) if paisagem else A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=20 * mm,
        title="LosPrice", author=_empresa(),
    )


def _estilo_tabela(cores_linha=None, alinhamentos=None):
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import TableStyle

    comandos = [
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(CINZA_CLARO)),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor(CINZA)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 1), (-1, -1), HexColor(PRETO)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, HexColor(LARANJA)),
        ("LINEBELOW", (0, 1), (-1, -2), 0.25, HexColor("#E3E6EB")),
    ]
    for coluna, alinhamento in (alinhamentos or {}).items():
        comandos.append(("ALIGN", (coluna, 0), (coluna, -1), alinhamento))
    for linha, cor in (cores_linha or {}).items():
        comandos.append(("TEXTCOLOR", (0, linha), (-1, linha), HexColor(cor)))
    return TableStyle(comandos)


# ---------------------------------------------------------------------------
# PDF - ficha tecnica
# ---------------------------------------------------------------------------


def ficha_tecnica_pdf(receita_id, caminho=None):
    """Ficha para imprimir e colar na cozinha."""
    from reportlab.lib.colors import HexColor
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table

    from controllers import receitas as rec_ctrl

    receita = rec_ctrl.obter(receita_id)
    if not receita:
        raise ValueError("Receita nao encontrada.")

    resultado = rec_ctrl.calcular(receita["itens"], receita["rendimento"])
    caminho = caminho or _nome_arquivo(
        "ficha_" + receita["nome"].lower().replace(" ", "_"), "pdf")

    estilos = _estilos()
    elementos = []

    _cabecalho(
        elementos, estilos, receita["nome"],
        f"Ficha tecnica  ·  {receita.get('categoria') or 'Sem categoria'}  ·  "
        f"rende {receita['rendimento']:g} {receita['unidade_rend'].lower()}"
        + (f"  ·  preparo {receita['tempo_preparo']} min"
           if receita.get("tempo_preparo") else ""))

    # resumo de custo
    resumo = [
        ["CUSTO DA RECEITA", "CUSTO POR UNIDADE", "INGREDIENTES", "EMBALAGENS"],
        [formatar_moeda(resultado.custo_total),
         formatar_moeda(resultado.custo_unitario),
         formatar_moeda(resultado.custo_ingredientes),
         formatar_moeda(resultado.custo_embalagens)],
    ]
    tabela = Table(resumo, colWidths=[43 * mm] * 4)
    tabela.setStyle(_estilo_tabela())
    tabela.setStyle(_estilo_tabela({1: LARANJA}))
    elementos.append(tabela)

    # itens
    elementos.append(Paragraph("Composicao", estilos["secao"]))

    linhas = [["ITEM", "QUANTIDADE", "CUSTO", "PESO"]]
    for detalhe in resultado.detalhes:
        unidade = detalhe["unidade"]
        linhas.append([
            detalhe["nome"] + ("  (embalagem)" if detalhe["tipo"] == "embalagem" else ""),
            f"{detalhe['quantidade']:g} {unidade}",
            formatar_moeda(detalhe["custo"]),
            formatar_pct(detalhe["participacao_pct"]),
        ])
    linhas.append(["TOTAL", "", formatar_moeda(resultado.custo_total), "100,0%"])

    tabela = Table(linhas, colWidths=[80 * mm, 32 * mm, 30 * mm, 30 * mm])
    tabela.setStyle(_estilo_tabela(alinhamentos={1: "RIGHT", 2: "RIGHT", 3: "RIGHT"}))
    tabela.setStyle(_estilo_tabela({len(linhas) - 1: LARANJA},
                                   {1: "RIGHT", 2: "RIGHT", 3: "RIGHT"}))
    elementos.append(tabela)

    # precos
    with conectar() as cur:
        cur.execute(
            """
            SELECT c.nome, p.preco_praticado, p.lucro_liquido, p.margem_real_pct
              FROM precificacao p
              JOIN canais c ON c.id = p.canal_id
             WHERE p.receita_id = ? AND c.ativo = 1
             ORDER BY c.ordem, c.id
            """,
            (receita_id,),
        )
        precos = [dict(l) for l in cur.fetchall()]

    if precos:
        elementos.append(Paragraph("Preco por canal", estilos["secao"]))
        linhas = [["CANAL", "PRECO", "LUCRO", "MARGEM"]]
        cores = {}
        for indice, preco in enumerate(precos, start=1):
            linhas.append([
                preco["nome"],
                formatar_moeda(preco["preco_praticado"]),
                formatar_moeda(preco["lucro_liquido"]),
                formatar_pct(preco["margem_real_pct"]),
            ])
            if preco["margem_real_pct"] < 0:
                cores[indice] = VERMELHO

        tabela = Table(linhas, colWidths=[80 * mm, 32 * mm, 30 * mm, 30 * mm])
        tabela.setStyle(_estilo_tabela(cores, {1: "RIGHT", 2: "RIGHT", 3: "RIGHT"}))
        elementos.append(tabela)

    # modo de preparo
    if receita.get("modo_preparo"):
        elementos.append(Paragraph("Modo de preparo", estilos["secao"]))
        elementos.append(Paragraph(receita["modo_preparo"].replace("\n", "<br/>"),
                                   estilos["corpo"]))

    elementos.append(Spacer(1, 10 * mm))
    elementos.append(Paragraph(
        "Respeitar as gramaturas acima e o que mantem o custo real igual ao "
        "custo calculado. Qualquer alteracao na ficha muda o preco de venda.",
        estilos["nota"]))

    _documento(caminho).build(elementos, onFirstPage=_rodape, onLaterPages=_rodape)
    return caminho


# ---------------------------------------------------------------------------
# PDF - tabela de precos
# ---------------------------------------------------------------------------


def tabela_precos_pdf(caminho=None):
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Table

    caminho = caminho or _nome_arquivo("tabela_precos", "pdf")

    with conectar() as cur:
        cur.execute("SELECT id, nome FROM canais WHERE ativo = 1 ORDER BY ordem, id")
        canais = [dict(l) for l in cur.fetchall()]

        cur.execute(
            """
            SELECT r.id, r.nome, r.custo_unitario
              FROM receitas r
             WHERE r.ativo = 1
               AND EXISTS (SELECT 1 FROM precificacao p WHERE p.receita_id = r.id)
             ORDER BY r.nome COLLATE NOCASE
            """
        )
        receitas = [dict(l) for l in cur.fetchall()]

        cur.execute("SELECT receita_id, canal_id, preco_praticado FROM precificacao")
        precos = {(l["receita_id"], l["canal_id"]): l["preco_praticado"]
                  for l in cur.fetchall()}

    if not receitas:
        raise ValueError("Nenhum produto precificado ainda.")

    estilos = _estilos()
    elementos = []
    _cabecalho(elementos, estilos, "Tabela de precos",
               f"{len(receitas)} produtos  ·  {len(canais)} canais de venda")

    cabecalho = ["PRODUTO", "CUSTO"] + [c["nome"].upper() for c in canais]
    linhas = [cabecalho]

    for receita in receitas:
        linha = [receita["nome"], formatar_moeda(receita["custo_unitario"])]
        for canal in canais:
            valor = precos.get((receita["id"], canal["id"]))
            linha.append(formatar_moeda(valor) if valor else "--")
        linhas.append(linha)

    largura_canal = (250 - 60 - 24) / max(len(canais), 1)
    larguras = [60 * mm, 24 * mm] + [largura_canal * mm] * len(canais)

    alinhamentos = {i: "RIGHT" for i in range(1, len(cabecalho))}
    tabela = Table(linhas, colWidths=larguras, repeatRows=1)
    tabela.setStyle(_estilo_tabela(alinhamentos=alinhamentos))
    elementos.append(tabela)

    elementos.append(Paragraph(
        "Precos calculados pelo metodo divisor: comissao, cartao e imposto "
        "incidem sobre o preco de venda, nao sobre o custo.",
        estilos["nota"]))

    _documento(caminho, paisagem=True).build(
        elementos, onFirstPage=_rodape, onLaterPages=_rodape)
    return caminho


# ---------------------------------------------------------------------------
# PDF - engenharia de cardapio
# ---------------------------------------------------------------------------


def engenharia_pdf(caminho=None):
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table

    dados = engenharia()
    if not dados["pronto"]:
        raise ValueError("Informe quantas unidades cada produto vende por mes.")

    caminho = caminho or _nome_arquivo("engenharia_cardapio", "pdf")

    estilos = _estilos()
    elementos = []
    _cabecalho(
        elementos, estilos, "Engenharia de cardapio",
        f"margem media {formatar_pct(dados['margem_media'])}  ·  "
        f"corte de popularidade em {dados['vendas_corte']:.0f} vendas/mes  ·  "
        f"lucro estimado {formatar_moeda(dados['lucro_mes_total'])}/mes")

    linhas = [["PRODUTO", "VENDE/MES", "MARGEM", "LUCRO/UN", "LUCRO/MES", "CLASSE"]]
    cores = {}
    for indice, produto in enumerate(dados["classificados"], start=1):
        linhas.append([
            produto["nome"],
            f"{produto['vendas_mes']:.0f}",
            formatar_pct(produto["margem"]),
            formatar_moeda(produto["lucro"]),
            formatar_moeda(produto["lucro_mes"]),
            produto["classe"],
        ])
        cores[indice] = produto["classe_cor"]

    tabela = Table(linhas,
                   colWidths=[52 * mm, 22 * mm, 22 * mm, 24 * mm, 26 * mm, 28 * mm],
                   repeatRows=1)
    tabela.setStyle(_estilo_tabela(cores, {1: "RIGHT", 2: "RIGHT",
                                           3: "RIGHT", 4: "RIGHT"}))
    elementos.append(tabela)

    elementos.append(Paragraph("Como ler", estilos["secao"]))
    guia = [["CLASSE", "SIGNIFICA", "O QUE FAZER"]]
    for classe, (texto, acao, _cor) in CLASSIFICACAO.items():
        guia.append([classe, texto, acao])

    tabela = Table(guia, colWidths=[26 * mm, 55 * mm, 93 * mm])
    tabela.setStyle(_estilo_tabela())
    elementos.append(tabela)

    elementos.append(Spacer(1, 6 * mm))
    elementos.append(Paragraph(
        "A classificacao compara cada produto com a media da sua propria casa: "
        "acima ou abaixo da margem media e do volume medio.",
        estilos["nota"]))

    _documento(caminho).build(elementos, onFirstPage=_rodape, onLaterPages=_rodape)
    return caminho


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _estilizar_planilha(aba, larguras):
    from openpyxl.styles import Alignment, Font, PatternFill

    fundo = PatternFill("solid", fgColor="FFF0E5")
    fonte = Font(bold=True, color="C25100", size=10)

    for celula in aba[1]:
        celula.fill = fundo
        celula.font = fonte
        celula.alignment = Alignment(vertical="center")

    for coluna, largura in zip("ABCDEFGHIJKLMNO", larguras):
        aba.column_dimensions[coluna].width = largura

    aba.freeze_panes = "A2"


def exportar_excel(caminho=None):
    """Uma planilha com quatro abas: insumos, embalagens, fichas e precos."""
    from openpyxl import Workbook

    caminho = caminho or _nome_arquivo("losprice", "xlsx")
    livro = Workbook()

    # --- ingredientes -----------------------------------------------------
    aba = livro.active
    aba.title = "Ingredientes"
    aba.append(["Ingrediente", "Categoria", "Fornecedor", "Qtd comprada",
                "Unidade", "Valor pago", "Fator", "Custo por unidade", "Base"])

    with conectar() as cur:
        cur.execute(
            """
            SELECT i.*, f.nome AS fornecedor
              FROM ingredientes i
              LEFT JOIN fornecedores f ON f.id = i.fornecedor_id
             WHERE i.ativo = 1 ORDER BY i.nome COLLATE NOCASE
            """
        )
        for linha in cur.fetchall():
            aba.append([
                linha["nome"], linha["categoria"] or "", linha["fornecedor"] or "",
                linha["qtd_comprada"], linha["unidade_compra"], linha["valor_pago"],
                linha["fator_correcao"], round(linha["custo_unitario"], 6),
                ROTULO_BASE.get(linha["unidade_base"], ""),
            ])
    _estilizar_planilha(aba, [30, 20, 24, 14, 10, 13, 8, 18, 8])

    # --- embalagens -------------------------------------------------------
    aba = livro.create_sheet("Embalagens")
    aba.append(["Embalagem", "Tipo", "Fornecedor", "Qtd comprada",
                "Valor pago", "Custo por unidade"])

    with conectar() as cur:
        cur.execute(
            """
            SELECT e.*, f.nome AS fornecedor
              FROM embalagens e
              LEFT JOIN fornecedores f ON f.id = e.fornecedor_id
             WHERE e.ativo = 1 ORDER BY e.nome COLLATE NOCASE
            """
        )
        for linha in cur.fetchall():
            aba.append([linha["nome"], linha["tipo"] or "", linha["fornecedor"] or "",
                        linha["qtd_comprada"], linha["valor_pago"],
                        round(linha["custo_unitario"], 4)])
    _estilizar_planilha(aba, [30, 16, 24, 14, 13, 18])

    # --- fichas tecnicas --------------------------------------------------
    aba = livro.create_sheet("Fichas tecnicas")
    aba.append(["Receita", "Item", "Tipo", "Quantidade", "Unidade", "Custo"])

    with conectar() as cur:
        cur.execute(
            """
            SELECT r.nome AS receita, i.nome AS item, ri.quantidade,
                   ri.unidade, ri.custo_calc
              FROM receita_ingredientes ri
              JOIN receitas r     ON r.id = ri.receita_id
              JOIN ingredientes i ON i.id = ri.ingrediente_id
             WHERE r.ativo = 1
             ORDER BY r.nome COLLATE NOCASE, ri.id
            """
        )
        for linha in cur.fetchall():
            aba.append([linha["receita"], linha["item"], "Ingrediente",
                        linha["quantidade"], linha["unidade"],
                        round(linha["custo_calc"], 4)])

        cur.execute(
            """
            SELECT r.nome AS receita, e.nome AS item, re.quantidade, re.custo_calc
              FROM receita_embalagens re
              JOIN receitas r   ON r.id = re.receita_id
              JOIN embalagens e ON e.id = re.embalagem_id
             WHERE r.ativo = 1
             ORDER BY r.nome COLLATE NOCASE, re.id
            """
        )
        for linha in cur.fetchall():
            aba.append([linha["receita"], linha["item"], "Embalagem",
                        linha["quantidade"], "UN", round(linha["custo_calc"], 4)])
    _estilizar_planilha(aba, [30, 30, 14, 12, 10, 12])

    # --- precos -----------------------------------------------------------
    aba = livro.create_sheet("Precos")
    aba.append(["Produto", "Canal", "Custo", "Preco sugerido", "Preco praticado",
                "Lucro", "Margem %"])

    with conectar() as cur:
        cur.execute(
            """
            SELECT r.nome AS receita, c.nome AS canal, r.custo_unitario,
                   p.preco_sugerido, p.preco_praticado, p.lucro_liquido,
                   p.margem_real_pct
              FROM precificacao p
              JOIN receitas r ON r.id = p.receita_id
              JOIN canais   c ON c.id = p.canal_id
             WHERE r.ativo = 1
             ORDER BY r.nome COLLATE NOCASE, c.ordem
            """
        )
        for linha in cur.fetchall():
            aba.append([
                linha["receita"], linha["canal"], round(linha["custo_unitario"], 2),
                round(linha["preco_sugerido"], 2),
                round(linha["preco_praticado"] or 0, 2),
                round(linha["lucro_liquido"], 2),
                round(linha["margem_real_pct"], 2),
            ])
    _estilizar_planilha(aba, [30, 28, 12, 16, 16, 12, 12])

    livro.save(caminho)
    return caminho


# ---------------------------------------------------------------------------
# Listagem
# ---------------------------------------------------------------------------


def receitas_disponiveis():
    with conectar() as cur:
        cur.execute(
            "SELECT id, nome, categoria, custo_unitario FROM receitas "
            " WHERE ativo = 1 ORDER BY nome COLLATE NOCASE"
        )
        return [dict(l) for l in cur.fetchall()]


def gerados():
    """Arquivos ja gerados, do mais recente para o mais antigo."""
    garantir_pasta()
    arquivos = []
    for nome in os.listdir(PASTA_RELATORIOS):
        caminho = os.path.join(PASTA_RELATORIOS, nome)
        if not os.path.isfile(caminho):
            continue
        info = os.stat(caminho)
        arquivos.append({
            "nome": nome,
            "caminho": caminho,
            "tamanho": info.st_size,
            "data": info.st_mtime,
            "tipo": "PDF" if nome.lower().endswith(".pdf") else "Excel",
        })
    return sorted(arquivos, key=lambda a: a["data"], reverse=True)[:20]
